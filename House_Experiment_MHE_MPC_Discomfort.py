import requests
from   nordpool import elspot, elbas
import json
from   datetime import timezone
from   datetime import date, datetime, timedelta
import pandas as pd
import sched, time
import pytz
import tzlocal
import matplotlib.pyplot as plt
import numpy as np
import scipy.interpolate as scinterp
import sched, time
import sensibo_client as SC
import tibber
from solcast.base import Base
from urllib.parse import urljoin
from isodate import parse_datetime, parse_duration
import sys
import pickle
import copy
import csv
import codecs
import urllib.request
import numpy, scipy.io
import math
from bs4 import BeautifulSoup
from metno_locationforecast import Place, Forecast
sys.path.append('/Applications/casadiPython3')
from casadi import *
from casadi.tools import *
import openpyxl
import string
 
 
# TODO:
#
# Make a user-friendly display.
#
# For each room:
# - Show temperature plan (desired and measured only) + on/off
# - Show elec. prices vs. power consumption plan
# - Show last 24h and next 24h
# - Display the marginal costs? (multipliers associated to the desired and min temperatures)
#
# For the house:
# - Show overall consumption + prices
# - Show last 24h and next 24h

plt.close("all")

# API keys collected here
APIkey = {'Sensibo' : 'tXZ5Ibk5s95UZgZgOU3wzJ4MOunl2f',
          'Tibber'  : 'dx61b7ne2HzGbpt0gI0xrB1f0jxd0_heXs-dmoMcnuA',
          'Weather' : 'c512438f8c9099db879c8fec88b10111',
          'Weather2': 'bafb0de79fd54009a0d205751210901',
          'Solcast' : 'hjIrt_Z00zbp4aoF7meqeH_YRwTtWBLk'
}

# If Raspberry needs a restart:
# - Go to local network
# - ssh pi@raspberrypi.local type usual password
# - (docker ps to see who's there)
# - cd raspberry-sw
# - docker-compose down (stops software)
# - docker-compose up -d (starts sotftware, run in the background)


# Raspberry addresses if Global is outdated, do:
# - Connect to local network
# - Go to http://raspberrypi.local:4551/status
# - Copy URL



#PIIP = 'http://192.168.1.44'               # Local
#PIIP = 'http://raspberry.local'
PIIP = 'https://3604010d2957.ngrok.io'      # Global (not operational yet)



#home_name = 'home-up'
home_name = 'home-pumps'

Tax = 0.25   # Tax on spot market prices

#### Setup Heat Pumps structure ####

global HPstates, HPmeas, HPNameList
HPstates   = ['targetTemperature','fanLevel','on','mode']
HPmeas     = ['temperature','humidity']
HPNameList = ['Entrance         ',
              'Living           ',
              'Studio           ',
              'Room Downstairs  ']

FanSpeeds = ['quiet','low','medium','medium_high','high']

# Heat Pumps colors
ColorList = ['b','g','c','m']

# Initialize Heat Pump drivers
try:
    # Create sensibo home object
    HomeSensibo = SC.SensiboClientAPI(APIkey['Sensibo'])
    # Get heat pumps (Sensibo objects)
    SensiboDevices = HomeSensibo.devices()
    print('Sensibo Devices :')
    print('-----------------')
    unit = 0
    Colors  = {}
    HPNames = {}
    for Device in SensiboDevices.keys():
        print(Device)
        Colors[Device]  = ColorList[unit]
        HPNames[Device] = HPNameList[unit]
        unit += 1
except:
    print('Home Sensibo could not be accessed (check network?). Code terminated')
    sys.exit()

Pumps = list(SensiboDevices.keys())
Npumps = len(Pumps)

ColHP = {}
for index, HP in enumerate(Pumps):
    ColHP[HP] = ColorList[index]



############################################
#                                          #
#  This code is testing MHE + MPC online  #
#                                          #
############################################



##########   Control parameters   ##########

# OPEN QUESTIONS
# - How to incorporate data efficiently in the MPC formulation?
# - Consumption at target = 16 deg is not zero, yields wrong power forecasts
#       1. try mixed-integer
#       2. can we use multiplers to decide on on/off?

# MHE-MPC meta parameters
MHE_Horizon  = 24 # in hours
MPC_Horizon  = 24 # in hours

# Some basic data
GridCost     = 44
MaxPowSingle = 1.5

# Parameters for on/off decisions
SwitchOnOffWindow = 30      #in minutes: minimum window of time for which an on/off decision is held
PowerSwitch       = 50/1e3  #in kW     : mean predicted power (over window) at which a switch on/off is prescribed

SpotBasePriceWindow = 1 # DEPRACATED - How many days in the past do we look at to calculate base prices (zero-gradient prices)

# Mixed integer formulation
MixInteger     = False
IntegerHorizon = 1

HPGroups = [
            ['living'],
            [ 'main',  'livingdown', 'studio' ]
          ]

MaxPowGroup = [1.8,2.5]

TargetTempLimit = {
                    'main'       : {'Min' : 10, 'Max' : 28},
                    'living'     : {'Min' : 10, 'Max' : 28},
                    'studio'     : {'Min' : 10, 'Max' : 28},
                    'livingdown' : {'Min' : 10, 'Max' : 28}
                   }

"""
HPControl = {}
for pump in Pumps:
    HPControl[pump] = {  'WarmTime'   : [6,23],
                         'WarmTemp'   : 21,
                         'ColdTemp'   : 19,
                         'Price_Gain' : .1,
                         'temp_max'   : 26}

HPControl['main']['WarmTime']         = [7,22]
HPControl['main']['WarmTemp' ]        = 21
HPControl['main']['ColdTemp' ]        = 18
HPControl['main']['Price_Gain']       = .12

HPControl['living']['WarmTime']       = [7,22]
HPControl['living']['WarmTemp' ]      = 21
HPControl['living']['ColdTemp' ]      = 18
HPControl['living']['Price_Gain']     = .1

HPControl['studio']['WarmTime']       = [19,8]
HPControl['studio']['WarmTemp' ]      = 20
HPControl['studio']['ColdTemp' ]      = 17
HPControl['studio']['Price_Gain']     = .13

HPControl['livingdown']['WarmTime']   = [20,8]
HPControl['livingdown']['WarmTemp' ]  = 20
HPControl['livingdown']['ColdTemp' ]  = 17
HPControl['livingdown']['Price_Gain'] = .13
"""
#####################################################


############### Prepare API stuff ###################



SYSIDFile = 'SYSID_Data_2021-03-05_9_SYSID_II'

print('Use SYSID file : '+SYSIDFile)

f = open(SYSIDFile+'.pkl',"rb")
SYSID = pickle.load(f)
f.close()

#sys.exit()
### Define stuff

# Measurement sampling time
SamplingTime = {'Measurement' :  5}   # Sampling time for HP

N_MHE_Horizon = int(60*MHE_Horizon/SamplingTime['Measurement'])
N_MPC_Horizon = int(60*MPC_Horizon/SamplingTime['Measurement'])




# Create scheduler for time management
s = sched.scheduler(time.time, time.sleep)

# Initialize Nordpool spot prices
prices_spot = elspot.Prices(currency='NOK')

# Position & Time zone
lat  = '63.4'                                  # Latitude    : Byasen
lon  = '10.335'                                # Longitude   : Byasen
Zone = 'Tr.heim'                               # Spot Market : Trondheim
local_timezone=pytz.timezone('Europe/Oslo')    # Time zone   : Oslo



# Met NO weather
location_name = 'Ugla'
lat_num       = 63.3998
long_num      = 10.3355
alti          = 200

HomeWeather          =  Place(location_name,lat_num,long_num,alti)
PersonalID           =  'sebastien.gros@ntnu.no POWIOT project'
HomeWeather_forecast = Forecast(HomeWeather, 'sebastien.gros@ntnu.no POWIOT project', forecast_type='compact')


# A dummy function for managing time
def dummy():
    ###
    return



##################  Useful functions for IoT and stuff #########################
def ReadSettings():
    AList = ['A','B','C','D','E','F','G']

    ## Read desired temperatures
    settings = openpyxl.load_workbook('TempSettings.xlsx')
    settings = settings.active

    TempSettings = {}
    for column in settings.iter_cols(1, settings.max_column):
        TempSettings[column[0].value] = []
       
    for i, row in enumerate(settings.iter_rows(values_only=True)):
        if i > 0:
            for colnum, colname in enumerate(TempSettings.keys()):
                TempSettings[colname].append(row[colnum])

    ## Read minimum temperatures
    settings = openpyxl.load_workbook('MinTempSettings.xlsx')
    settings = settings.active

    MinTempSettings = {}
    for column in settings.iter_cols(1, settings.max_column):
        MinTempSettings[column[0].value] = []
       
    for i, row in enumerate(settings.iter_rows(values_only=True)):
        if i > 0:
            for colnum, colname in enumerate(MinTempSettings.keys()):
                MinTempSettings[colname].append(row[colnum])

    ## Read MHE & MPC weights
    settings = openpyxl.load_workbook('Weights.xlsx')
    settings = settings.active

    ColNames = []
    for column in settings.iter_cols(1, settings.max_column):
        ColNames.append(column[0].value)

    indexMPC = ColNames.index('MPC')
    indexMHE = ColNames.index('MHE')
    indexExt = ColNames.index('External')

    WeightMPC = {}
    WeightMHE = {}
    WeightExt = {}
    for i, row in enumerate(settings.iter_rows(values_only=True)):
        if i > 0 and not(row[indexMPC]==None) and not(row[indexMPC+1]==None):
            WeightMPC[row[indexMPC]] = row[indexMPC+1]
        if i > 0 and not(row[indexMHE]==None) and not(row[indexMHE+1]==None):
            WeightMHE[row[indexMHE]] = row[indexMHE+1]
        if i > 0 and not(row[indexExt]==None) and not(row[indexExt+1]==None):
            WeightExt[row[indexExt]] = row[indexExt+1]

    return TempSettings, MinTempSettings, WeightMPC, WeightMHE, WeightExt




def PiFirstCall(TimeInitial, extension='all'):

    print('#############################')
    print('#     Pull Raspberry Pi     #')
    print('#############################')
    #print('Time initial : '+str(TimeInitial))
    print('Extension    : '+extension)

    print('####### Pull sensibo ########')
    
    DataHP     = CreatePumpDataStructures(SensiboDevices)             # Empty structure to dump data (state & measurements)

    response = requests.get(PIIP+'/api/get_data/sensibo/daily/'+extension)
    Sensibo = json.loads(response.text)

    dates = list(Sensibo['data'].keys())
    for pump in Pumps:
        #print(pump)

        # If empty structure, fill it
        for date in dates:
            #print(date)
            codes = list(Sensibo['data'][date].keys())
            for code in codes:
                #print('Code : '+str(code))
                for time in Sensibo['data'][date][code][pump]['time']:
                    time_local = datetime.strptime(time,'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)
                    DataHP[pump]['flattime'].append((time_local-TimeInitial).total_seconds()/60.)
                    DataHP[pump]['Time'].append(time_local)
                for type in ['measurements','states']:
                    for item in Sensibo['data'][date][code][pump][type].keys():
                        DataHP[pump][type][item] += Sensibo['data'][date][code][pump][type][item]
                #print('--------------------------------------')
                    
                    
    print('#######  Pull Tibber  #######')
    response = requests.get(PIIP+'/api/get_data/tibber-realtime-'+home_name+'/daily/'+extension)
    TibberRT = json.loads(response.text)

    TibberData = {'Time'      : [],
                  'flattime'  : [],
                  'Power'     : [],
                  'Energy'    : [] }
              
    # Extract Tibber data
    dates = list(TibberRT['data'].keys())


    # First call, data structure is empty, fill it
    for date in dates:
        #print(date)
        codes = list(TibberRT['data'][date].keys())
        for code in codes:
            #print('Code : '+str(code))
            #print('--------------------------------------')
            for index, time in enumerate(TibberRT['data'][date][code]['time']):
                    time_local = datetime.strptime(time,'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)#+timedelta(hours=+1)
                    TibberData['Time'].append(time_local)
                    TibberData['flattime'].append((time_local-TimeInitial).total_seconds()/60.)
                    TibberData['Power'].append(TibberRT['data'][date][code]['power'][index])
                    TibberData['Energy'].append(TibberRT['data'][date][code]['accumulatedConsumption'][index])

    return DataHP, TibberData


def PiCall(DataHP, TibberData, TimeInitial, extension='all'):

    print('#############################')
    print('#     Pull Raspberry Pi     #')
    print('#############################')
    #print('Time initial : '+str(TimeInitial))
    print('Extension    : '+extension)
    
    print('####### Pull sensibo ########')
    
    try:
        response = requests.get(PIIP+'/api/get_data/sensibo/daily/'+extension, timeout=60)
        Sensibo = json.loads(response.text)

        dates = list(Sensibo['data'].keys())
        
        NewHPTime = {}
        for pump in Pumps:
            #print(pump)
            NewHPTime[pump] = []
            for date in dates:
                #print(date)
                codes = list(Sensibo['data'][date].keys())
                for code in codes:
                    #print('Code : '+str(code))
                    Index = -1
                    NewTime = datetime.strptime(Sensibo['data'][date][code][pump]['time'][-1],'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)#+timedelta(hours=-1)
                    while NewTime > DataHP[pump]['Time'][-1] and len(Sensibo['data'][date][code][pump]['time'])+Index > 0:
                        Index -= 1
                        NewTime = datetime.strptime(Sensibo['data'][date][code][pump]['time'][Index],'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)#+timedelta(hours=-1)

                    for index in range(Index+1,0):
                        time       = Sensibo['data'][date][code][pump]['time'][index]
                        time_local = datetime.strptime(time,'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)
                        #print('TimeInitial : '+str(TimeInitial))
                        #print('Time local  : '+str(time_local))
                        #print('Flat time   : '+str((time_local-TimeInitial).total_seconds()/60.))

                        NewHPTime[pump].append((time_local-TimeInitial).total_seconds()/60.)
                        DataHP[pump]['Time'].append(time_local)
                        DataHP[pump]['flattime'].append((time_local-TimeInitial).total_seconds()/60.)
                        for type in ['measurements','states']:
                            for item in Sensibo['data'][date][code][pump][type].keys():
                                DataHP[pump][type][item].append(Sensibo['data'][date][code][pump][type][item][index])
                    #print('--------------------------------------')
            #print('Latest '+str(pump)+' data : '+str(TibberData['Time'][-1]))
    except:
        print('Sensibo data pull failed')

                    
    print('#######  Pull Tibber  #######')
    TibberDataNew = {'Time'      : [],
                     'flattime'  : [],
                     'Power'     : [],
                     'Energy'    : [] }
                         
    try:
        response = requests.get(PIIP+'/api/get_data/tibber-realtime-'+home_name+'/daily/'+extension, timeout=60)
        TibberRT = json.loads(response.text)

        # Extract Tibber data
        dates = list(TibberRT['data'].keys())


        # Data structure is populated, add new data only
        for date in dates:
            codes = list(TibberRT['data'][date].keys())
            for code in codes:
                #print('Code : '+str(code))
                #print('--------------------------------------')
                Index = -1
                
                NewTime = datetime.strptime(TibberRT['data'][date][code]['time'][-1],'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)#+timedelta(hours=+1)
                while NewTime > TibberData['Time'][-1] and len(TibberRT['data'][date][code]['time'])+Index > 0:
                    Index -= 1
                    NewTime = datetime.strptime(TibberRT['data'][date][code]['time'][Index],'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)#+timedelta(hours=+1)
                
                for index in range(Index+1,0):
                        time       = TibberRT['data'][date][code]['time'][index]
                        time_local = datetime.strptime(time,'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)#+timedelta(hours=-1)
                        #print('TimeInitial : '+str(TimeInitial))
                        #print('Time local  : '+str(time_local))
                        #print('Flat time   : '+str((time_local-TimeInitial).total_seconds()/60.))
                        TibberData['Time'].append(time_local)
                        TibberData['flattime'].append((time_local-TimeInitial).total_seconds()/60.)
                        TibberData['Power'].append(TibberRT['data'][date][code]['power'][index])
                        TibberData['Energy'].append(TibberRT['data'][date][code]['accumulatedConsumption'][index])
                        
                        TibberDataNew['Time'].append(time_local)
                        TibberDataNew['flattime'].append((time_local-TimeInitial).total_seconds()/60.)
                        TibberDataNew['Power'].append(TibberRT['data'][date][code]['power'][index])
                        TibberDataNew['Energy'].append(TibberRT['data'][date][code]['accumulatedConsumption'][index])
        print('Latest Tibber data : '+str(TibberData['Time'][-1]))
    except:
        print('Tibber data pull failed')
        
    return DataHP, TibberData, TibberDataNew

def PiShortCall(DataHP, TibberData, TimeInitial, window=2*SamplingTime['Measurement']):

    print('#############################')
    print('#     Pull Raspberry Pi     #')
    print('#############################')
    #print('Time initial : '+str(TimeInitial))
    print('Window    : '+str(window)+'min')
    
    print('####### Pull sensibo ########')
    
    try:
        response = requests.get(PIIP+'/api/get_data_since/sensibo/minutes/'+str(window), timeout=60)
        Sensibo = json.loads(response.text)
        if Sensibo['data']:
            print('--------------------------------------')
            NewHPTime = {}
            for pump in Pumps:
                print(pump)
                print('Latest data at :'+str(DataHP[pump]['Time'][-1]))
                for time in Sensibo['data'][pump]['time']:
                    print('Data in window :'+str(datetime.strptime(time,'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)))
            
                NewHPTime[pump] = []
                Index = -1
                NewTime = datetime.strptime(Sensibo['data'][pump]['time'][-1],'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)#+timedelta(hours=-1)
                while NewTime > DataHP[pump]['Time'][-1] and len(Sensibo['data'][pump]['time'])+Index > 0:
                    Index -= 1
                    NewTime = datetime.strptime(Sensibo['data'][pump]['time'][Index],'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)#+timedelta(hours=-1)
                for index in range(Index,0):
                    time       = Sensibo['data'][pump]['time'][index]
                    time_local = datetime.strptime(time,'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)

                    if time_local > DataHP[pump]['Time'][-1]:
                        print('Add data at time '+str(time_local))
                        NewHPTime[pump].append((time_local-TimeInitial).total_seconds()/60.)
                        DataHP[pump]['Time'].append(time_local)
                        DataHP[pump]['flattime'].append((time_local-TimeInitial).total_seconds()/60.)
                        for type in ['measurements','states']:
                            for item in Sensibo['data'][pump][type].keys():
                                DataHP[pump][type][item].append(Sensibo['data'][pump][type][item][index])
                print('--------------------------------------')
        else:
            print('No Sensibo data in the '+str(window)+'min window')
    except:
        print('Sensibo data pull failed')

                    
    print('#######  Pull Tibber  #######')
    TibberDataNew = {'Time'      : [],
                     'flattime'  : [],
                     'Power'     : [],
                     'Energy'    : [] }
                         
    try:
        response = requests.get(PIIP+'/api/get_data_since/tibber-realtime-'+home_name+'/minutes/'+str(window), timeout=60)
        TibberRT = json.loads(response.text)


        # Data structure is populated, add new data only
        Index = -1
        
        NewTime = datetime.strptime(TibberRT['data']['time'][-1],'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)#+timedelta(hours=+1)
        while NewTime > TibberData['Time'][-1] and len(TibberRT['data']['time'])+Index > 0:
            Index -= 1
            NewTime = datetime.strptime(TibberRT['data']['time'][Index],'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)#+timedelta(hours=+1)
        
        for index in range(Index+1,0):
                time       = TibberRT['data']['time'][index]
                time_local = datetime.strptime(time,'%a, %d %b %Y %H:%M:%S %Z').replace(tzinfo=timezone.utc).astimezone(local_timezone)#+timedelta(hours=-1)
                #print('TimeInitial : '+str(TimeInitial))
                #print('Time local  : '+str(time_local))
                #print('Flat time   : '+str((time_local-TimeInitial).total_seconds()/60.))
                TibberData['Time'].append(time_local)
                TibberData['flattime'].append((time_local-TimeInitial).total_seconds()/60.)
                TibberData['Power'].append(TibberRT['data']['power'][index])
                TibberData['Energy'].append(TibberRT['data']['accumulatedConsumption'][index])
                
                TibberDataNew['Time'].append(time_local)
                TibberDataNew['flattime'].append((time_local-TimeInitial).total_seconds()/60.)
                TibberDataNew['Power'].append(TibberRT['data']['power'][index])
                TibberDataNew['Energy'].append(TibberRT['data']['accumulatedConsumption'][index])
        print('Latest Tibber data : '+str(TibberData['Time'][-1]))
    except:
        print('Tibber data pull failed')
        
    return DataHP, TibberData, TibberDataNew


def ExtractWeatherMetNo(Home_forecast,local_timezone):
    # More info at https://pypi.org/project/metno-locationforecast/
    Data = vars(Home_forecast.data)
    Weather = {'Time'        : [],
               'Temperature' : []}
               
    for data in Data['intervals']:
        Weather['Time'].append(  data.start_time.replace(tzinfo=timezone.utc).astimezone(local_timezone)  )
        Weather['Temperature'].append( data.variables['air_temperature'].value )
    
    return Weather
    
    
class RadiationForecasts(Base):

    end_point = 'radiation/forecasts'

    def __init__(self, latitude, longitude, *args, **kwargs):

        self.latitude = latitude
        self.longitude = longitude
        self.forecasts = None

        self.params = {'latitude' : self.latitude,
                       'longitude' : self.longitude}

        self._get(*args, **kwargs)

        if self.ok:
            self._generate_forecasts_dict()

    def _generate_forecasts_dict(self):

        self.forecasts = []

        for forecast in self.content.get('forecasts'):

            # Convert period_end and period. All other fields should already be
            # the correct type
            forecast['period_end'] = parse_datetime(forecast['period_end'])
            forecast['period'] = parse_duration(forecast['period'])

            self.forecasts.append(forecast)

def GetSpotMarketHistory(Zone,local_timezone,DaysBack):
    ########################  PULL SPOT MARKET  ######################################################
    try:
        Spot = {'Time': [],
                'Prices': []}
            
        DateList = []
        for days in range(-DaysBack,0):
            DateList.append( date.today() + timedelta(days=days) )
            
        for FetchDate in DateList:
            print('Fetch spot data: '+str(FetchDate))
            prices = prices_spot.hourly(areas=[Zone],end_date=FetchDate)

            for i in range(24):
                # Get spot market time stamp, local time
                time = prices['areas'][Zone]['values'][i]['start']        # UTC times
                time = time.astimezone(local_timezone)                    # convert to local time
                price  = prices['areas'][Zone]['values'][i]['value']/1e1  # convert to Øre/kWh
                price += Tax*price                                        # add VAT
                
                #print(str(time)+' '+str(price))

                
                Spot['Time'].append(time)
                Spot['Prices'].append(price)
    except:
        Spot = []

    return Spot

def GetSpotMarket(Zone,local_timezone):
    ########################  PULL SPOT MARKET  ######################################################
    try:
        Spot = {'Time': [],
                'Prices': []}

        # Find the noon of today (local timezone), needed to know if next day is avaiable
        timeNow   = datetime.now(tz=local_timezone)
        Noon      = timeNow.replace(second=0, microsecond=0, minute=0, hour=12) #Noon of the day (when spot prices are announced for the next 24h)

        # Pull the spot prices for yesterday (-1), today (0), and if afternoon for tomorrow (1)
        if timeNow > Noon:
            DayList    = [0,1]
        else:
            DayList    = [0]
            

        for day in DayList:
            #print(day)
            FetchDate = date.today() + timedelta(days=day)
            prices = prices_spot.hourly(areas=[Zone],end_date=FetchDate)
            
            for i in range(24):
                #print(str(prices['areas'][Zone]['values'][i]['value']))
                # Get spot market time stamp, local time
                time = prices['areas'][Zone]['values'][i]['start']        # UTC times
                time = time.astimezone(local_timezone)                    # convert to local time
                price = prices['areas'][Zone]['values'][i]['value']/1e1   # convert to Øre/kWh
                price += Tax*price                                        # add VAT

                Spot['Time'].append(time)
                Spot['Prices'].append(price)
    except:
        Spot = []

    return Spot

def Wait(TimeSchedule):
    TimeNow = datetime.now(tz=local_timezone)
    Delay = (TimeSchedule - TimeNow).total_seconds()
    if Delay > 0:
        s.enter(Delay,0,dummy)
        s.run()

def CreatePumpState(SensiboDevices):
   StateNames = {'targetTemperature': [], 'fanLevel': 'medium','on' : True}
   State = {}
   for HP in SensiboDevices.keys():
       CurrentState = HomeSensibo.pod_ac_state(SensiboDevices[HP])
       State[HP] = {}
       for key in StateNames.keys():
            if StateNames[key]:
                State[HP][key] = StateNames[key]
            else:
                State[HP][key] = CurrentState[key]

   return State
   
def CreatePumpDataStructures(SensiboDevices):
    # Creates structures with initial lists
    Data = {}
    for key in SensiboDevices.keys():
        Data[key]  =  {'flattime'      : [],
                       'Time'          : [],
                       'measurements'  : {},
                       'states'        : {}}
                       
        for keymeas in HPmeas:
            Data[key]['measurements'][keymeas] = []
        for keystate in HPstates:
            Data[key]['states'][keystate] = []
                        
    return Data

def CheckHPStates(SensiboDevices):
    for HP in SensiboDevices.keys():
        CurrentState = HomeSensibo.pod_ac_state(SensiboDevices[HP])
        if CurrentState['on'] == False and CurrentState['targetTemperature'] > 16:
            print('State '+HP+' corrected')
            HomeSensibo.pod_change_ac_state(SensiboDevices[HP],CurrentState,'on',True)
            HomeSensibo.pod_change_ac_state(SensiboDevices[HP],CurrentState,'targetTemperature',16)
                        
def UpdatePumpStates(SensiboDevices,State):
    # Updates the state of all pumps according to the data structure "Control"
    for HP in SensiboDevices.keys():
        print('----------------------------------------')
        print('Unit: '+str(HPNames[HP]))
        print('----------------------------------------')
        try:
            CurrentState = HomeSensibo.pod_ac_state(SensiboDevices[HP])
            
            # Check on/off state
            key = 'on'
            if CurrentState[key] == State[HP][key]:
                print('Current state '+str(key)+' : '+str(CurrentState[key]))
            
            # If pump off and must switch on, do that first
            if (CurrentState[key] == False) and (State[HP][key] == True):
                print('Current state '+str(key)+' : '+str(CurrentState[key])+' --> '+str(State[HP][key]))
                CurrentState[key] = State[HP][key]
                HomeSensibo.pod_change_ac_state(SensiboDevices[HP],CurrentState,key,State[HP][key])
                CurrentState[key] = State[HP][key]
                
            # Update all the other states
            if CurrentState['on'] == True:
                for key in State[HP].keys():
                    if not(key == 'on'):
                        if not(CurrentState[key] == State[HP][key]):
                            print('Current state '+str(key)+' : '+str(CurrentState[key])+' --> '+str(State[HP][key]))
                            HomeSensibo.pod_change_ac_state(SensiboDevices[HP],CurrentState,key,State[HP][key])
                            CurrentState[key] = State[HP][key]
                        else:
                            print('Current state '+str(key)+' : '+str(CurrentState[key]))
                        
            # If pump on and must switch off, do that last
            key = 'on'
            if (CurrentState[key] == True) and (State[HP][key] == False):
                print('Current state '+str(key)+' : '+str(CurrentState[key])+' --> '+str(State[HP][key]))
                CurrentState[key] = State[HP][key]
                HomeSensibo.pod_change_ac_state(SensiboDevices[HP],CurrentState,key,State[HP][key])
                CurrentState[key] = State[HP][key]
        except:
            print(HP+' timeout, will try next sampling time')


def AvPower(time,power):
    # Compute averave power from time, power
    # time is in datetime format
    Energy = 0
    Dt     = 0
    for k in range(len(time)-1):
        dt      = (time[k+1] - time[k]).total_seconds()
        Dt     += dt
        Energy += power[k]*dt/1e3
    if Dt > 0:
        PowerAverage = Energy/Dt
    else:
        PowerAverage = power[-1]/1e3
    
    return PowerAverage
  
def AvPowerFlat(time,power,time_start=[],time_end=[]):
    # Compute averave power from time, power, from the end to time_star
    # time is in flat format
    
    if time:
        # If no time_start assigned or time start is below the time window:
        if not(time_start):
            time_start = time[0]
        else:
            if time_start < time[0]:
                time_start = time[0]
        
        # If no time_end assigned or time end is above the time window:
        if not(time_end):
            time_end = time[-1]
        else:
            if time_end > time[-1]:
                time_end = time[-1]

        dt     = 2. #seconds
        AvGrid = [time_start + dt/60.]
        while AvGrid[-1] < time_end:
                AvGrid.append( AvGrid[-1] + dt/60. )
        
        f            = scinterp.interp1d(np.array(time), np.array(power), kind='nearest', fill_value='extrapolate')
        PowerInterp  = list(f(np.array(AvGrid)))
        PowerAverage = np.mean(PowerInterp)/1e3
    else:
            PowerAverage = []
    return PowerAverage



print('Pull spot market')

Spot = GetSpotMarket(Zone,local_timezone)

print('Read settings')
TempSettings, MinTempSettings, WeightMPC, WeightMHE, WeightExt = ReadSettings()

WeightsLog = {'Times' :[],
              'MHE'   : {},
              'MPC'   : {}}


####################################################
################     BUILD MHE   ###################
####################################################

######## CREATE BASIC STRUCTURES ########
print('############## Build controller ##############')
print('- Build MHE')

DataSample = struct_symMX([ entry('Meas_temp'),
                            entry('Ref_temp'),
                            entry('On'),
                            entry('Out_temp'),
                            entry('Power')     ])

PumpParams = struct_symMX([ entry('Gain'),
                            entry('Const'),
                            entry('Room_loss'),
                            entry('Room_Gain'),
                            entry('WallGainIn'),
                            entry('WallGainOut')  ])
                            
MHEstates = struct_symMX([  entry('Temp'),
                            entry('WallTemp'),
                            entry('Power')      ])
                        
MHEinputs = struct_symMX([  entry('Power_Perturbation'),
                            entry('Temp_Perturbation')])

################ PREPARE MHE  ################

########################  Create parameters structures  ########################

Parameters = []
for pump in Pumps:
    Parameters += [entry( pump, struct = PumpParams)]

Parameters = struct_symMX( Parameters  )

########################  Create Data strutures  ########################
N = N_MHE_Horizon

DataSample = struct_symMX([ entry('Meas_temp'),
                            entry('Ref_temp'),
                            entry('On'),
                            entry('Out_temp')  ])

PumpsData = []
for pump in Pumps:
    PumpsData += [entry( pump, struct = DataSample)]
    
PumpsData += [entry( 'Power' )]
PumpsData  = struct_symMX( PumpsData )

WeightStruct = []
for key in WeightMHE.keys():
    WeightStruct+= [entry(key)]
    WeightsLog['MHE'][key] = []
WeightStruct = struct_symMX(WeightStruct)

DataMHE = struct_symMX([
                          entry( 'Data',           struct = PumpsData, repeat = N  ),
                          entry( 'Weights',        struct = WeightStruct  ),
                          entry( 'PrevParameters', struct = Parameters  )
                         ])


########################  Create dynamic variable structures  ########################

states_all = []
for pump in Pumps:
    states_all += [ entry( pump, struct = MHEstates) ]
states_all = struct_symMX( states_all  )

inputs_all = []
for pump in Pumps:
    inputs_all += [ entry( pump, struct = MHEinputs) ]
inputs_all = struct_symMX( inputs_all  )

########################  BUILD MHE   ########################

# Decision variables
wMHE = struct_symMX( [
                    entry('State', struct = states_all, repeat = N),
                    entry('Input', struct = inputs_all, repeat = N-1),
                    entry('Param', struct = Parameters)
                  ])

ReLu = 30.

J = 0

g      = []
lbgMHE = []
ubgMHE = []

lbwMHE = wMHE(-inf)
ubwMHE = wMHE(+inf)
wMHE0  = wMHE(0)
    
# Formulate the NLP

for k in range(N-1):
    ElecPower = 0
    for pump in Pumps:
        ### State dynamics ###
        DTWallIn   = wMHE['State',k,pump,'WallTemp']  - wMHE['State',k,pump,'Temp']
        DTWallOut  = DataMHE['Data',k,pump,'Out_temp'] - wMHE['State',k,pump,'WallTemp']
        
        DTtarget      = DataMHE['Data',k,pump,'Ref_temp'] - DataMHE['Data',k,pump,'Meas_temp']#wMHE['State',k,pump,'Temp']
        DTtarget_plus = DataMHE['Data',k+1,pump,'Ref_temp'] - DataMHE['Data',k,pump,'Meas_temp']#wMHE['State',k+1,pump,'Temp']
        
        Delta_Power         = wMHE['Param',pump,'Gain']*( DTtarget_plus - DTtarget)/10.  # From power model
        ElecPower_pump_plus = wMHE['State',k,pump,'Power'] + Delta_Power + wMHE['Input',k,pump,'Power_Perturbation']/1e3
        ElecPower_pump_plus = log(1+exp(ReLu*(  ElecPower_pump_plus  ) ) )/ReLu
        ElecPower_pump_plus = MaxPowSingle - log(1+exp(ReLu*(  MaxPowSingle - ElecPower_pump_plus  ) ) )/ReLu
        
    
        ElecPower  += DataMHE['Data',k,pump,'On']*wMHE['State',k,pump,'Power']
        
        COP        = 4.
        PumpPower  = DataMHE['Data',k,pump,'On']*COP*wMHE['State',k,pump,'Power']
        
        # Temperature dynamics, inputs_k[pump,'pow'] is the thermal power injected in the room: pumps + disturbances
        Delta_Room_Temp =  wMHE['Param',pump,'Room_loss']*DTWallIn/100. + wMHE['Param',pump,'Room_Gain']*PumpPower/1e2 + wMHE['Input',k,pump,'Temp_Perturbation']
        Delta_Wall_Temp = -wMHE['Param',pump,'WallGainIn']*DTWallIn/1e2 + wMHE['Param',pump,'WallGainOut']*DTWallOut/1e3
        
        g.append(  wMHE['State',k,pump,'Temp']     + Delta_Room_Temp  -  wMHE['State',k+1,pump,'Temp']     )
        lbgMHE.append(0)
        ubgMHE.append(0)
        
        g.append(  wMHE['State',k,pump,'WallTemp'] + Delta_Wall_Temp  -  wMHE['State',k+1,pump,'WallTemp']    )
        lbgMHE.append(0)
        ubgMHE.append(0)
        
        g.append(  ElecPower_pump_plus                                -  wMHE['State',k+1,pump,'Power']    )
        lbgMHE.append(0)
        ubgMHE.append(0)

        J +=               DataMHE['Weights','Temp']*(wMHE['State',k,pump,'Temp'] - DataMHE['Data',k,pump,'Meas_temp']      )**2/float(N*Npumps)
        J +=           DataMHE['Weights','WallTemp']*(        Delta_Wall_Temp                                               )**2/float(N*Npumps)
        J += DataMHE['Weights','Power_Perturbation']*(    wMHE['Input',k,pump,'Power_Perturbation']                         )**2/float(N*Npumps)   # Minimize Perturbation
        J +=  DataMHE['Weights','Temp_Perturbation']*(    wMHE['Input',k,pump,'Temp_Perturbation']                          )**2/float(N*Npumps)   # Minimize Perturbation
                
    J += DataMHE['Weights','Power']*len(Pumps)*(ElecPower -  DataMHE['Data',k,'Power'])**2

for k in range(N):
    for indexgroup, group in enumerate(HPGroups):
        PowerGroup = 0
        for pump in group:
            PowerGroup += DataMHE['Data',k,pump,'On']*wMHE['State',k,pump,'Power']
            
        g.append( PowerGroup )
        lbgMHE.append(-inf)
        ubgMHE.append( MaxPowGroup[indexgroup] )

for pump in Pumps:
    J += DataMHE['Weights','Temp']*(wMHE['State',-1,pump,'Temp'] - DataMHE['Data',-1,pump,'Meas_temp'])**2/float(Npumps)

for pump in Pumps:
    for key in SYSID[pump].keys():
        J += DataMHE['Weights','Param']*( wMHE['Param',pump,key] - DataMHE[ 'PrevParameters', pump, key ]  )**2

# Create an NLP solver
MHE = {'f': J, 'x': wMHE, 'g': vertcat(*g), 'p': DataMHE}

options = {}
options['ipopt'] = {'print_level':0}
solverMHE = nlpsol('solver', 'ipopt', MHE, options)

DataMHENum = DataMHE(0)  # Create matching numerical structure

SYSIDLog = {'Time' : []}
for pump in Pumps:
    SYSIDLog[pump] = {}
    for key in SYSID[pump].keys():
        SYSIDLog[pump][key]  = []

##############################################################
########################  BUILD MPC   ########################
##############################################################

print('- Build MPC')

########################  Create Data strutures  ########################
N = N_MPC_Horizon

TempRef = []
for pump in Pumps:
    TempRef += [entry( pump )]
TempRef = struct_symMX( TempRef  )

TempMin = []
for pump in Pumps:
    TempMin += [entry( pump )]
TempMin = struct_symMX( TempMin  )

MPCstates = struct_symMX([      entry('Temp'),
                                entry('TargetTemp'),
                                entry('WallTemp'),
                                entry('Power'),
                                entry('Discomfort'),
                                entry('SlackMinTemp')          ])


MPCinputs = struct_symMX([      entry('Delta_TargetTemp'),
                                entry('On'),
                                entry('Slack')   ])


DataMPC = [         entry('Out_temp',           repeat = N ),
                    entry('SpotPrices',         repeat = N ),
                    entry('DesiredTemperature', struct = TempRef, repeat = N ),
                    entry('MinTemperature',     struct = TempMin, repeat = N ),
                    entry('BasePrice')   ]

WeightStruct = []
for key in WeightMPC.keys():
    WeightStruct+= [entry(key)]
    WeightsLog['MPC'][key] = []
WeightStruct = struct_symMX(WeightStruct)

DataMPC += [ entry('Weights', struct=WeightStruct) ]

SYSIDParamPump = []
for key in SYSID[Pumps[0]].keys():
    SYSIDParamPump += [ entry(key) ]
SYSIDParamPump = struct_symMX(SYSIDParamPump)

SYSIDParam = []
for pump in Pumps:
    SYSIDParam += [ entry(pump, struct=SYSIDParamPump) ]
SYSIDParam = struct_symMX(SYSIDParam)
  
DataMPC += [ entry('SYSID', struct=SYSIDParam) ]
DataMPC = struct_symMX(DataMPC)


########################  Create dynamic variable structures  ########################

states_all = []
for pump in Pumps:
    states_all += [ entry( pump, struct = MPCstates) ]
states_all = struct_symMX( states_all  )

inputs_all = []
for pump in Pumps:
    inputs_all += [ entry( pump, struct = MPCinputs) ]
inputs_all = struct_symMX( inputs_all  )

# Decision variables
wMPC = struct_symMX( [
                    entry('State', struct = states_all, repeat = N),
                    entry('Input', struct = inputs_all, repeat = N-1),
                  ])
                  
J   = 0
g   = []

lbwMPC = wMPC(-inf)
ubwMPC = wMPC(+inf)
wMPC0  = wMPC(0)

lbgMPC = []
ubgMPC = []

# Formulate the NLP
for k in range(N-1):
    ElecPower = 0
    for pump in Pumps:
        ### State dynamics ###
        DTWallIn   = wMPC['State',k,pump,'WallTemp']  - wMPC['State',k,pump,'Temp']
        DTWallOut  = DataMPC['Out_temp',k] - wMPC['State',k,pump,'WallTemp']
                
        # THIS MODEL NEEDS TO BE REVISED FOR GROUP POWER!!!!
        DeltaTarget         = wMPC['Input',k,pump,'Delta_TargetTemp']
        Delta_Power         = DataMPC['SYSID',pump,'Gain']*( DeltaTarget - wMPC['State',k+1,pump,'Temp'] + wMPC['State',k,pump,'Temp'] )/10.  # From power model
        ElecPower_pump_plus = wMPC['State',k,pump,'Power'] + Delta_Power
            
        ElecPower  += wMPC['Input',k,pump,'On']*wMPC['State',k,pump,'Power']
        
        COP        = 4.
        PumpPower  = wMPC['Input',k,pump,'On']*COP*wMPC['State',k,pump,'Power']
        
        # Temperature dynamics, inputs_k[pump,'pow'] is the thermal power injected in the room: pumps + disturbances
        Delta_Room_Temp =  DataMPC['SYSID',pump,'Room_loss']*DTWallIn/100. + DataMPC['SYSID',pump,'Room_Gain']*PumpPower/1e2
        Delta_Wall_Temp = -DataMPC['SYSID',pump,'WallGainIn']*DTWallIn/1e2 + DataMPC['SYSID',pump,'WallGainOut']*DTWallOut/1e3

        g.append(  wMPC['State',k,pump,'TargetTemp'] + DeltaTarget  -  wMPC['State',k+1,pump,'TargetTemp']     )
        lbgMPC.append(0)
        ubgMPC.append(0)

        g.append(  wMPC['State',k,pump,'Temp']     + Delta_Room_Temp  -  wMPC['State',k+1,pump,'Temp']     )
        lbgMPC.append(0)
        ubgMPC.append(0)

        g.append(  wMPC['State',k,pump,'WallTemp'] + Delta_Wall_Temp  -  wMPC['State',k+1,pump,'WallTemp']    )
        lbgMPC.append(0)
        ubgMPC.append(0)
        
        g.append(  ElecPower_pump_plus                                -  wMPC['State',k+1,pump,'Power']    )
        lbgMPC.append(0)
        ubgMPC.append(0)

        g.append( DataMPC['Weights','DiscomfortDecay']*wMPC['State',k,pump,'Discomfort'] + DataMPC['Weights','DiscomfortGain']*wMPC['Input',k,pump,'Slack'] - wMPC['State',k+1,pump,'Discomfort'] )
        lbgMPC.append(0)
        ubgMPC.append(0)

        g.append(  DataMPC['DesiredTemperature',k,pump] - wMPC['State',k,pump,'Temp'] - wMPC['Input',k,pump,'Slack']   )
        lbgMPC.append(-inf)
        ubgMPC.append(0)

        g.append(  DataMPC['MinTemperature',k,pump] - wMPC['State',k,pump,'Temp'] - wMPC['State',k,pump,'SlackMinTemp']   )
        lbgMPC.append(-inf)
        ubgMPC.append(0)
    
        ### Cost ###
        HubDeltaTemp   = DataMPC['Weights','HubDeltaTemp']
        HubDeltaTarget = DataMPC['Weights','HubDeltaTarget']
        SlackPenalty   = (HubDeltaTemp**2)*( sqrt( 1 + (wMPC['Input',k,pump,'Slack']/HubDeltaTemp)**2) - 1 )
        
        J += DataMPC['Weights','SpotGain']*( DataMPC['SpotPrices',k] + GridCost - DataMPC['BasePrice'] )*ElecPower/float(N)
        J += DataMPC['Weights','TempBelow']*wMPC['State',k,pump,'Discomfort']**2/float(N)
        J += DataMPC['Weights','TempAbove']*(DataMPC['DesiredTemperature',k,pump] - wMPC['State',k,pump,'Temp'])**2/float(N)
        J += DataMPC['Weights','MinTemp']*wMPC['State',k,pump,'SlackMinTemp']**2/float(N)
        
        J += DataMPC['Weights','Delta_TargetTemp']*(HubDeltaTarget**2)*(sqrt( 1 + (wMPC['Input',k,pump,'Delta_TargetTemp']/HubDeltaTarget)**2 ) - 1)/float(N)
        
        
    if k > 0:
        for indexgroup, group in enumerate(HPGroups):
            PowerGroup = 0
            for pump in group:
                PowerGroup += wMPC['Input',k,pump,'On']*wMPC['State',k,pump,'Power']
                
            g.append( PowerGroup )
            lbgMPC.append(-inf)
            ubgMPC.append( MaxPowGroup[indexgroup] )

    for pump in Pumps:
        J += DataMPC['Weights','TempBelow']*wMPC['State',-1,pump,'Discomfort']**2/float(N)

        g.append(  DataMPC['MinTemperature',-1,pump] - wMPC['State',-1,pump,'Temp'] - wMPC['State',-1,pump,'SlackMinTemp']   )
        lbgMPC.append(-inf)
        ubgMPC.append(0)
        J += DataMPC['Weights','MinTemp']*wMPC['State',-1,pump,'SlackMinTemp']**2/float(N)

# Create an NLP solver
MPC = {'f': J, 'x': wMPC, 'g': vertcat(*g), 'p': DataMPC}
options['ipopt'] = {'print_level':0}

if MixInteger:
    discrete = wMPC(False)
    for pump in Pumps:
        discrete['Input',:IntegerHorizon,pump,'Delta_TargetTemp_Up']   = True
        discrete['Input',:IntegerHorizon,pump,'Delta_TargetTemp_Down'] = True
    discrete = [bool(item) for item in np.array(discrete.cat)]

    solverMPC = nlpsol('solver', 'bonmin', MPC, {'discrete': discrete})
else:
    solverMPC = nlpsol('solver', 'ipopt', MPC, options)

DataMPCNum = DataMPC(0)

############## Prepare figures ##############

print('############## Prepare plotting ##############')

plt.ion()
FigList = []
AxList  = []
Ax2List = []
for fig in range(0,17):
    figID = plt.figure(fig+1,figsize=(18,6))
    FigList.append(figID)
    AxList.append( [] )
    Ax2List.append( [] )
    for sub in range(1,3):
        ax = figID.add_subplot(1,2,sub)
        AxList[fig].append(ax)
        Ax2List[fig].append(ax.twinx())


for fig in range(17,18):
    figID = plt.figure(fig+1,figsize=(18,6))
    FigList.append(figID)
    AxList.append( [] )
    for sub in range(1,7):
        ax = figID.add_subplot(3,2,sub)
        AxList[fig].append(ax)


############## Prepare MHE/MPC logger ##############

Logger_MHE_MPC = {'MHE' : {'DateTime'    : [],
                           'Time_Grid'   : [],
                           'RawSolution' : []},
                  'MPC' : {'DateTime'    : [],
                           'Time_Grid'   : [],
                           'RawSolution' : []}
                    }

print('Pull weather')
HomeWeather_forecast.update()
Forecast   = ExtractWeatherMetNo(HomeWeather_forecast,local_timezone)
NextUpdate = vars(HomeWeather_forecast.data)['expires'].replace(tzinfo=timezone.utc).astimezone(local_timezone)

print('############## Prepare Data structures ##############')

        
HPState     = CreatePumpState(SensiboDevices)

Weather = {'Time'        : [ Forecast['Time'][0] ],
           'Temperature' :  [ Forecast['Temperature'][0] ]}


SpotMarket = {'Time'     :  Spot['Time'],
              'Prices'   :  Spot['Prices'],
              'flattime' : [] }


PowerAverage ={'Times': [],
               'Power': []}


MHELog = {'Times' : [],
          'Times_datetime' : [],
          'Power' : []}
            
for pump in Pumps:
    MHELog[pump] = {'State' : {},
                    'Input' : {}}
    for key in MHEstates.keys():
        MHELog[pump]['State'][key] = []
    for key in MHEinputs.keys():
        MHELog[pump]['Input'][key] = []

    
############# Set schedule ############################


TimeSchedule = datetime.now(tz=local_timezone)
dt = SamplingTime['Measurement']
if int(dt*np.ceil(TimeSchedule.minute/dt)) < 60:
    TimeSchedule = TimeSchedule.replace(second=0,microsecond=0,minute=int(dt*np.ceil(TimeSchedule.minute/dt)))
else:
    if TimeSchedule.hour < 23:
        TimeSchedule = TimeSchedule.replace(second=0,microsecond=0,minute=0,hour=TimeSchedule.hour+1)
    else:
        TimeSchedule = TimeSchedule.replace(second=0,microsecond=0,minute=0,hour=0,day=TimeSchedule.day+1)

TimeInitial   = TimeSchedule
TimeFinal     = TimeInitial  + timedelta(hours=24*5)

for time in Spot['Time']:
    SpotMarket['flattime'].append( (time - TimeInitial).total_seconds()/60. )


FileName = 'MHE_MPC_Experiments_'+str(TimeInitial.date())+'_'+str(TimeInitial.hour)+'_'+str(TimeInitial.minute)


OnOffSwitch = {}
for pump in Pumps:
    OnOffSwitch[pump] = TimeSchedule - timedelta(hours=1)
    
## Data name

################## Pull Pi data (first call) ##################################

dates = str(date.today()-timedelta(days=1))
DataHP, TibberData = PiFirstCall(TimeInitial, extension = dates)
dates = str(date.today())
DataHP, TibberData, TibberDataNew  = PiCall(DataHP, TibberData , TimeInitial, extension = dates)


##################  Prepare data structures #####################3

print('Data dump in : '+FileName+'.pkl' )


Times = {'Weather'   : [],
         'Energy'    : [],
         'Spot'      : [],
         'Forecast'  : [],
         'Radiation' : []}

Radiation = {'Time'  : [],
             'GHI'   : [],
             'DNI'   : [],
             'DHI'   : [],
             'ELE'   : [],
             'AZI'   : [],
             'GTI'   : [],
             'Theta' : [],
             'Temp'  : []
}


Times['Weather'].append( (Weather['Time'][-1] - TimeInitial).total_seconds()/60.  )

PowerLog      = {'Times' : [],
                 'Power' : []}

RadiationData = {'Time' : [],
                 'DNI'   : []}



# Flat time grid for the MHE-MPC schemes
time_grid     = [ int(SamplingTime['Measurement']*np.ceil(TibberData['flattime'][0]/float(SamplingTime['Measurement']))) ]
PowerAverage  = [ TibberData['Power'][0] ]

while time_grid[-1] < -SamplingTime['Measurement']:
    time_grid.append( time_grid[-1] +  SamplingTime['Measurement'])
    PowerAverage.append(AvPowerFlat(TibberData['flattime'],TibberData['Power'],time_start = time_grid[-2] ,time_end = time_grid[-1]))

MHEWarmStarted         = False
MPCWarmStarted         = False
MPCControlAvailable    = False
MHEEstimationAvailable = False



print('---------- Start ------------')
####### RUN THE CODE ##########
while 0 <= 1:
    print('##########################################################################################################')
    print('Next time sample : '+str(TimeSchedule))
    print('##########################################################################################################')
    Wait(TimeSchedule)

    time_grid.append((TimeSchedule - TimeInitial).total_seconds()/60.)

    #################### Read settings ####################
    print('Read Settings')
    TempSettings, MinTempSettings, WeightMPC, WeightMHE, WeightExt = ReadSettings()
    
    WeightsLog['Times'].append(TimeSchedule)
    for key in WeightMHE.keys():
        WeightsLog['MHE'][key].append(WeightMHE[key])
    for key in WeightMPC.keys():
        WeightsLog['MPC'][key].append(WeightMPC[key])


    #################### Read spot market ####################
    # If spot market horizon less than 12h, try to pull
    try:
        if (Spot['Time'][-1] - TimeSchedule).total_seconds()/3600. < 10 or not(np.isfinite(Spot['Prices'][-1])) :
            print('Pull spot market data')
            SpotNew = GetSpotMarket(Zone,local_timezone)
            if SpotNew['Prices'][-1] < np.inf:
                Spot = SpotNew
                for k, time in enumerate(SpotNew['Time']):
                    if time > SpotMarket['Time'][-1]:
                        SpotMarket['Prices'].append(SpotNew['Prices'][k])
                        SpotMarket['Time'].append(time)
                        SpotMarket['flattime'].append( (time - TimeInitial).total_seconds()/60. )

    except:
        print('Exception in Spot prices')

                         
    # Make sure that the prices are finite
    for index, price in enumerate(Spot['Prices']):
        if not(price < np.inf):
            Spot['Prices'][index] = Spot['Prices'][index-1]

    Times['spot'] = []
    for time in Spot['Time']:
        Times['spot'].append( (time - TimeInitial).total_seconds()/60. )
    
    
    #################### Read Weather ####################
    # If weather data are obsolete, update
    if TimeSchedule > NextUpdate:
        try:
            print('Pull weather data')
            HomeWeather_forecast.update()
            Forecast   = ExtractWeatherMetNo(HomeWeather_forecast,local_timezone)
            NextUpdate = vars(HomeWeather_forecast.data)['expires'].replace(tzinfo=timezone.utc).astimezone(local_timezone)
            if Forecast['Time'][0] > Weather['Time'][-1]:
                Weather['Time'].append(Forecast['Time'][0])
                Weather['Temperature'].append(Forecast['Temperature'][0])
                Times['Weather'].append( (Forecast['Time'][0] - TimeInitial).total_seconds()/60.  )
            else:
                if not( Forecast['Temperature'][0] == Weather['Temperature'][-1] ):
                    Weather['Temperature'][-1] = Forecast['Temperature'][0]
        except:
            print('Exception in Weather')
            
    Times['Forecast'] = []
    for time in Forecast['Time']:
        Times['Forecast'].append( (time - TimeInitial).total_seconds()/60. )

    #################### Read Raspberry Pi ####################
    #DataHP, TibberData, TibberDataNew = PiCall(DataHP, TibberData , TimeInitial, extension = 'latest')
    
    # Check window that should be pulled:
    OldestData = TibberData['Time'][-1]
    for pump in Pumps:
        if DataHP[pump]['Time'][-1] < OldestData:
            OldestData = DataHP[pump]['Time'][-1]
    window = np.int(np.ceil((TimeSchedule - OldestData).total_seconds()/60.))
    DataHP, TibberData, TibberDataNew  = PiShortCall(DataHP, TibberData , TimeInitial , window=window)
    
    # Process power average
    NewPowerAverage = AvPowerFlat(TibberDataNew['flattime'],TibberDataNew['Power'],time_start = time_grid[-2] ,time_end = time_grid[-1])
    if NewPowerAverage:
        PowerAverage.append(NewPowerAverage)
    else:
        PowerAverage.append(PowerAverage[-1])
        print('WARNING: POWER DATA MISSING')
    
    """
    # If radiation data are more than 12h old, try to pull
    try:
        Update = False
        if not(Radiation['Time']):
            # If Radiation not pulled yet, do it
            Update = True
        else:
            if (TimeSchedule - RadiationData['Time'][-1]).total_seconds()/3600. >= .5:
                # If Radiation 12h or older then pull
                Update = True
            
        if Update:
            print('Pull radiations')
            RadForcast = RadiationForecasts(lat_num,long_num,api_key=APIkey['Solcast'])
            Radiation = {'Time'  : [],
                         'GHI'   : [],
                         'DNI'   : [],
                         'DHI'   : [],
                         'ELE'   : [],
                         'AZI'   : [],
                         'GTI'   : [],
                         'Theta' : [],
                         'Temp'  : []}
            for data in RadForcast.forecasts:
                Radiation['Time'].append(data['period_end'].astimezone(local_timezone)-data['period'])
                Radiation['GHI'].append(data['ghi'])
                Radiation['DNI'].append(data['dni'])
                Radiation['DHI'].append(data['dhi'])
                Radiation['ELE'].append(data['zenith'])    # Angle to the vertical
                Radiation['AZI'].append(data['azimuth'])   # Angle from the south
                Radiation['Temp'].append(data['air_temp'])   # Angle from the south

            # Append first 24h of the forecast in data
            time_index = 0
            while Radiation['Time'][time_index] < Radiation['Time'][0] + timedelta(hours=12):
                RadiationData['Time'].append(Radiation['Time'][time_index])
                RadiationData['DNI'].append(Radiation['DNI'][time_index])
                Times['Radiation'].append( (Radiation['Time'][time_index] - TimeInitial).total_seconds()/60. )
                time_index += 1
    except:
        print('Exception in Radiation')
    """
    
    #### Prepare time grids for MHE and MPC
    MHE_time_grid = time_grid[-N_MHE_Horizon:]
    MPC_time_grid = [ time_grid[-1] + k*SamplingTime['Measurement'] for k in range(0,N)]

    # Complete the time grid if data window is not long enough
    MHE_time_grid = [0.]*(N_MHE_Horizon-len(MHE_time_grid)) + MHE_time_grid

    #######################################################################
    ########################  Prepare & Solve MHE  ########################
    #######################################################################

    # MHE uses: Times, DataHP, Weather, PowerAverage
        
    for pump in Pumps:
        DataMHENum['Data',:,pump,'Meas_temp'] = list(np.interp(  MHE_time_grid, DataHP[pump]['flattime'], DataHP[pump]['measurements']['temperature']    ))
        DataMHENum['Data',:,pump,'Ref_temp']  = list(np.interp(  MHE_time_grid, DataHP[pump]['flattime'], DataHP[pump]['states']['targetTemperature']    ))
        DataMHENum['Data',:,pump,'On']        = list(np.round(np.interp(  MHE_time_grid, DataHP[pump]['flattime'], DataHP[pump]['states']['on']          )))
        DataMHENum['Data',:,pump,'Out_temp']  = list(np.interp(  MHE_time_grid, Times['Weather'], Weather['Temperature']                ))
    DataMHENum['Data',:,'Power'] = list(np.interp(  MHE_time_grid, time_grid, PowerAverage                                     ))
   
    for key in WeightMHE:
        DataMHENum['Weights',key] = WeightMHE[key]
    
    if not(MHEWarmStarted):
        # Initial guesses and bounds
        for pump in Pumps:
            wMHE0['State',:,pump,'WallTemp']         =  DataMHENum['Data',:,pump,'Meas_temp']
            lbwMHE['State',:,pump,'WallTemp']        =  DataMHENum['Data',:,pump,'Out_temp']
            ubwMHE['State',:,pump,'WallTemp']        =  [1.25*val for val in DataMHENum['Data',:,pump,'Meas_temp']]

            wMHE0['State',:,pump,'Power']            =  .75
            lbwMHE['State',:,pump,'Power']           =  0
            ubwMHE['State',:,pump,'Power']           =  1.5


            wMHE0['State',:,pump,'Temp']             =  DataMHENum['Data',:,pump,'Meas_temp']
            
            #lbwMHE['Input',:,pump,'Perturbation']    = -10
            #ubwMHE['Input',:,pump,'Perturbation']    = +10

    # Initial guesses and bounds for model parameters
    fac = WeightExt['Fac']
    if MHEWarmStarted:
        for pump in Pumps:
            for key in SYSID[pump].keys():
                wMHE0['Param',pump,key]  = SYSID[pump][key] #wMHE_opt['Param',pump,key]
                lbwMHE['Param',pump,key] = (1-fac)*SYSID[pump][key] #wMHE_opt['Param',pump,key]
                ubwMHE['Param',pump,key] = (1+fac)*SYSID[pump][key] #wMHE_opt['Param',pump,key]
                DataMHENum[ 'PrevParameters', pump, key ] = SYSID[pump][key] #wMHE_opt['Param',pump,key]
    else:
        for pump in Pumps:
            for key in SYSID[pump].keys():
                wMHE0['Param',pump,key]  = SYSID[pump][key]
                lbwMHE['Param',pump,key] = (1-fac)*SYSID[pump][key]
                ubwMHE['Param',pump,key] = (1+fac)*SYSID[pump][key]
                DataMHENum[ 'PrevParameters', pump, key ] = SYSID[pump][key]

    # Solve the NLP
    sol      = solverMHE(x0=wMHE0, lbx=lbwMHE, ubx=ubwMHE, lbg=lbgMHE, ubg=ubgMHE, p=DataMHENum)
    wMHE_opt = sol['x'].full().flatten()
    
    Logger_MHE_MPC['MHE']['RawSolution'].append(wMHE_opt)
    Logger_MHE_MPC['MHE']['Time_Grid'].append(MHE_time_grid)
    Logger_MHE_MPC['MHE']['DateTime'].append(TimeSchedule)
    
    wMHE_opt = wMHE(wMHE_opt)
    if solverMHE.stats()['success']:
        MHEWarmStarted         = True
        MHEEstimationAvailable = True
    else:
        MHEWarmStarted         = False
        MHEEstimationAvailable = False

    if MHEEstimationAvailable:
        SYSIDLog['Time'].append((TimeSchedule - TimeInitial).total_seconds()/60.)
        for pump in Pumps:
            print(pump)
            print('--------------')
            for key in SYSID[pump].keys():
                print('Value '+key+' : '+str(wMHE_opt['Param',pump,key])+' | From Value : '+str(wMHE0['Param',pump,key])+' | SYSID : '+str(SYSID[pump][key]))
                SYSIDLog[pump][key].append(float(wMHE_opt['Param',pump,key]))
                
    if solverMHE.stats()['success']:
        wMHE0    = wMHE_opt

    MHEElecPower = []
    NpumpOn   = []
    for k in range(N_MHE_Horizon):
        MHEElecPower.append(0)
        NpumpOn.append(0)
        for pump in Pumps:
            MHEElecPower[-1]  += float(DataMHENum['Data',k,pump,'On']*wMHE_opt['State',k,pump,'Power'])
            NpumpOn[-1]       += DataMHENum['Data',k,pump,'On']
    
    
    MHELog['Times'].append(MHE_time_grid[-1])
    MHELog['Power'].append(MHEElecPower[-1])
    MHELog['Times_datetime'].append(TimeSchedule)
    for pump in Pumps:
        for key in MHEstates.keys():
            MHELog[pump]['State'][key].append(wMHE_opt['State',-1,pump,key])
        for key in MHEinputs.keys():
            MHELog[pump]['Input'][key].append(wMHE_opt['Input',-1,pump,key])


    FN = 0
    for k, pump in enumerate(Pumps):
        AxList[FN][0].clear()
        AxList[FN][1].clear()
        Ax2List[FN][0].clear()
        Ax2List[FN][1].clear()
        AxList[FN][0].step(DataHP[pump]['flattime'], DataHP[pump]['measurements']['temperature'],color='b')
        AxList[FN][0].step(DataHP[pump]['flattime'], DataHP[pump]['states']['targetTemperature'],color='b',linestyle=':')
        AxList[FN][0].step(MHE_time_grid,wMHE_opt['State',:,pump,'Temp'],color='r',linewidth=3)
        AxList[FN][0].step(MHE_time_grid,wMHE_opt['State',:,pump,'WallTemp'],color='k')

        AxList[FN][0].set(title=pump)
        AxList[FN][0].legend(['Meas. temp.','Target. temp.','Est. temp.','Est. Wall temp.'])
        AxList[FN][0].grid()
        
        AxList[FN][1].step(MHE_time_grid,wMHE_opt['State',:,pump,'Power'],color='b')
        Ax2List[FN][1].step(MHE_time_grid[:-1],wMHE_opt['Input',:,pump,'Temp_Perturbation'],color='r')
        AxList[FN][1].set_xlim(MHE_time_grid[0],MHE_time_grid[-1])
        Ax2List[FN][1].set_ylabel('Temp correction [deg/5min]',fontsize=20,color='r')
        AxList[FN][1].set_ylabel('Power [kwh]',fontsize=20,color='b')
        AxList[FN][1].set(title='HP Power')
        AxList[FN][1].grid()
        FigList[FN].canvas.draw()
        
        FN += 1
        
    AxList[FN][0].clear()
    AxList[FN][1].clear()
    AxList[FN][0].step(time_grid,PowerAverage,color='b',where='post')
    AxList[FN][0].step(MHE_time_grid,MHEElecPower,color='r',where='post')
    AxList[FN][0].set_ylim([0,4.5])
    AxList[FN][0].set(title='Total Pumps Power')
    AxList[FN][0].grid()
    Leg = []
    for pump in Pumps:
        Leg.append(pump)
        AxList[FN][1].step(MHE_time_grid[:-1],wMHE_opt['Input',:,pump,'Power_Perturbation'],color=ColHP[pump])
    AxList[FN][1].legend(Leg)
    AxList[FN][1].set(title='Power adjustments in [W]')
    AxList[FN][1].grid()
    FigList[FN].canvas.draw()
        
    #######################################################################
    ########################  Prepare & Solve MPC  ########################
    #######################################################################

    TimeTemp = []
    for k in range(N_MPC_Horizon):
        TimeAbs = TimeSchedule + timedelta(minutes=SamplingTime['Measurement']*k)
        TimeTemp.append( np.mod(TimeAbs.hour + TimeAbs.minute/60., 24))

    for pump in Pumps:
        
        lbwMPC['State',0,pump,'WallTemp']         =  wMHE_opt['State',-1,pump,'WallTemp']
        ubwMPC['State',0,pump,'WallTemp']         =  wMHE_opt['State',-1,pump,'WallTemp']

        lbwMPC['State',:,pump,'Power']            =  0
        ubwMPC['State',:,pump,'Power']            =  1.5
        lbwMPC['State',0,pump,'Power']            =  wMHE_opt['State',-1,pump,'Power']*DataMHENum['Data',-1,pump,'On']
        ubwMPC['State',0,pump,'Power']            =  wMHE_opt['State',-1,pump,'Power']*DataMHENum['Data',-1,pump,'On']

        lbwMPC['State',0,pump,'Temp']             =  wMHE_opt['State',-1,pump,'Temp']
        ubwMPC['State',0,pump,'Temp']             =  wMHE_opt['State',-1,pump,'Temp']

        lbwMPC['State',:,pump,'TargetTemp']       =  TargetTempLimit[pump]['Min']
        ubwMPC['State',:,pump,'TargetTemp']       =  TargetTempLimit[pump]['Max']
        
        lbwMPC['Input',:,pump,'Delta_TargetTemp'] =  -inf
        ubwMPC['Input',:,pump,'Delta_TargetTemp'] =  +inf
        
        lbwMPC['Input',:,pump,'On']               =  1
        ubwMPC['Input',:,pump,'On']               =  1

        lbwMPC['Input',0,pump,'On']               =  DataMHENum['Data',-1,pump,'On']
        ubwMPC['Input',0,pump,'On']               =  DataMHENum['Data',-1,pump,'On']

        lbwMPC['Input',:,pump,'Slack']            =  0
        ubwMPC['Input',:,pump,'Slack']            =  +inf
                
        lbwMPC['State',:,pump,'SlackMinTemp']     =  0
        ubwMPC['State',:,pump,'SlackMinTemp']     =  +inf

        if not(MPCWarmStarted):
            wMPC0['State',:,pump,'WallTemp']          =  wMHE_opt['State',-1,pump,'WallTemp']
            wMPC0['State',:,pump,'Power']             =  wMHE_opt['State',-1,pump,'Power']
            wMPC0['State',:,pump,'Temp']              =  wMHE_opt['State',-1,pump,'Temp']
            wMPC0['State',:,pump,'TargetTemp']        =  DataMHENum['Data',-1,pump,'Ref_temp']
            wMPC0['Input',:,pump,'Delta_TargetTemp']  =  0
            wMPC0['Input',:,pump,'On']                =  1
            
            wMPC0['Input',:,pump,'Slack']             =  0
            
            lbwMPC['State',0,pump,'TargetTemp']       =  DataMHENum['Data',-1,pump,'Ref_temp']
            ubwMPC['State',0,pump,'TargetTemp']       =  DataMHENum['Data',-1,pump,'Ref_temp']


            lbwMPC['State',0,pump,'Discomfort']       =  0
            ubwMPC['State',0,pump,'Discomfort']       =  0

        else:
            if wMPC0['State',1,pump,'TargetTemp'] < 16 and DataMHENum['Data',-1,pump,'On'] == False:
                lbwMPC['State',0,pump,'TargetTemp']   =  wMPC0['State',1,pump,'TargetTemp']
                ubwMPC['State',0,pump,'TargetTemp']   =  wMPC0['State',1,pump,'TargetTemp']
            else:
                lbwMPC['State',0,pump,'TargetTemp']   =  DataMHENum['Data',-1,pump,'Ref_temp']
                ubwMPC['State',0,pump,'TargetTemp']   =  DataMHENum['Data',-1,pump,'Ref_temp']
                            
            lbwMPC['State',0,pump,'Discomfort']       =  wMPC0['State',1,pump,'Discomfort']
            ubwMPC['State',0,pump,'Discomfort']       =  wMPC0['State',1,pump,'Discomfort']

            if DataMHENum['Data',-1,pump,'On'] == False:
                # If pump is off, assign 16 by default
                lbwMPC['State',0,pump,'TargetTemp']       =  np.min([wMPC0['State',1,pump,'TargetTemp'],16])
                ubwMPC['State',0,pump,'TargetTemp']       =  np.min([wMPC0['State',1,pump,'TargetTemp'],16])

        f = scinterp.interp1d(np.array(TempSettings['Times']), np.array(TempSettings[pump]), kind='nearest')
        DataMPCNum['DesiredTemperature',:,pump] = list(f(np.array(TimeTemp)))

        f = scinterp.interp1d(np.array(MinTempSettings['Times']), np.array(MinTempSettings[pump]), kind='nearest')
        DataMPCNum['MinTemperature',:,pump] = list(f(np.array(TimeTemp)))


    DataMPCNum['Out_temp',:]                  = list(np.interp(  MPC_time_grid,  Times['Forecast'], Forecast['Temperature']           ))
    
    
    #####  Feed spot market into MPC  #########
    # Create a piecewise constant spot signal for easy interpolation
    # Note: spot price applies from the time window starting at the corresponding time in the list up to +1h
    SpotTimes  = []
    SpotPrices = []
    for k, price in enumerate(Spot['Prices']):
        SpotTimes.append(Times['spot'][k])
        SpotTimes.append(Times['spot'][k]+60)
        SpotPrices.append(price)
        SpotPrices.append(price)
    
    f = scinterp.interp1d(np.array(SpotTimes), np.array(SpotPrices), kind='nearest', fill_value=Spot['Prices'][-1], bounds_error=False)
    DataMPCNum['SpotPrices',:]                = list(f(np.array(MPC_time_grid)))
            
    """
    plt.figure(21)
    plt.plot(Times['spot'],Spot['Prices'],color='k',linewidth=3)
    plt.step(Times['spot'],Spot['Prices'], where='post',color='k',linewidth=3)
    plt.plot(SpotTimes,SpotPrices,color='b',linewidth=2)
    plt.step(SpotMarket['flattime'],SpotMarket['Prices'], where='post',color='c')
    plt.plot(MPC_time_grid,np.array(DataMPCNum['SpotPrices',:]),color='r')
    plt.grid('on')
    """
    
    ######################
    #Base price
    DataMPCNum['BasePrice'] = 0
    print('Base Price : '+str(DataMPCNum['BasePrice'])+' Øre/kWh')
    
    for key in WeightMPC.keys():
        DataMPCNum['Weights',key] = WeightMPC[key]
        
    for pump in Pumps:
        for key in SYSID[pump].keys():
            DataMPCNum['SYSID',pump,key] = wMHE_opt['Param',pump,key]

    # Solve the NLP
    print('###############################################')
    print('###############    Solve MPC    ###############')
    print('###############################################')
    sol      = solverMPC(x0=wMPC0, lbx=lbwMPC, ubx=ubwMPC, lbg=lbgMPC, ubg=ubgMPC, p=DataMPCNum)
    wMPC_opt = sol['x'].full().flatten()
    
    Logger_MHE_MPC['MPC']['RawSolution'].append(wMPC_opt)
    Logger_MHE_MPC['MPC']['Time_Grid'].append(MPC_time_grid)
    Logger_MHE_MPC['MPC']['DateTime'].append(TimeSchedule)
    
    wMPC_opt = wMPC(wMPC_opt)

    if solverMPC.stats()['success']:
        wMPC0 = wMPC_opt
        MPCWarmStarted      = True
        MPCControlAvailable = True
    else:
        MPCWarmStarted      = False
        MPCControlAvailable = False
        
        for pump in Pumps:
            print(pump)
            print(str(lbwMPC['State',0,pump,'Power']))
            print(str(ubwMPC['State',0,pump,'Power']))
        
        print('MPC failed')
        sys.exit()

    MPCElecPower = []
    for k in range(N_MPC_Horizon-1):
        MPCElecPower.append(0)
        for pump in Pumps:
            MPCElecPower[-1]  += float(wMPC_opt['Input',k,pump,'On']*wMPC_opt['State',k,pump,'Power'])

    MPCActions = {}
    for pump in Pumps:
        MPCActions[pump] = {
                            'RoomTemp'   : float(wMPC_opt['State',0,pump,'Temp']),
                            'TargetTemp' : int(np.round(wMPC_opt['State',1,pump,'TargetTemp'])),
                            'On'         : bool(wMPC_opt['Input',1,pump,'On']),
                            'PowerMean'  : np.mean(wMPC_opt['State',1:1+int(np.round(SwitchOnOffWindow/float(SamplingTime['Measurement']))),pump,'Power'])
                            }
                            
    MPCControlAvailable = True
    print('------------ MPC Actions ------------')
    for pump in Pumps:
        print(pump+' : on '+str(MPCActions[pump]['On']))
        print('Target temperature : '+str(MPCActions[pump]['TargetTemp']))
     
     
    print('--------------------------')
    print('Weights MPC')
    print('-------------------------')
    for key in WeightMPC:
         print(key+' : '+str(WeightMPC[key]))
    print('--------------------------')
    print('Weights MHE')
    print('--------------------------')
    for key in WeightMHE:
         print(key+' : '+str(WeightMHE[key]))
    print('--------------------------')
    print('External parameters')
    print('--------------------------')
    for key in WeightExt:
         print(key+' : '+str(WeightExt[key]))
    print('--------------------------')
    
    ####### Implement Heat Pump control here ########

    # On/off switching
    for pump in Pumps:
        HPState[pump]['targetTemperature'] = int(np.max([16,MPCActions[pump]['TargetTemp']]))
        
        if (MPCActions[pump]['PowerMean'] > PowerSwitch) and (MPCActions[pump]['On'] == True) and HPState[pump]['on'] == False:
            if (TimeSchedule - OnOffSwitch[pump]).total_seconds() > SwitchOnOffWindow*60:
                #Switch on if the MPC pump power is high enough, and if the last switch is old enough
                HPState[pump]['on']                = True
                OnOffSwitch[pump]                  = TimeSchedule
                print(pump+' switch at : '+str(OnOffSwitch[pump]) )
        if (MPCActions[pump]['PowerMean'] < PowerSwitch) and HPState[pump]['on'] == True:
            if (TimeSchedule - OnOffSwitch[pump]).total_seconds() > SwitchOnOffWindow*60:
                HPState[pump]['on']                = False
                HPState[pump]['targetTemperature'] = 16
                OnOffSwitch[pump]                  = TimeSchedule
                print(pump+' switch at : '+str(OnOffSwitch[pump]) )
                
    #Assign Fan speed: Adopt higher fan speeds for larger temperature errors
    for pump in Pumps:
        HPState[pump]['fanLevel']     = 'medium'
        TempDiff = MPCActions[pump]['RoomTemp'] - DataHP[pump]['measurements']['temperature'][-1]
        TempTarg = HPState[pump]['targetTemperature'] - DataHP[pump]['measurements']['temperature'][-1]
        fanspeed   = np.max([np.min([np.max([0,3*TempDiff]),4]), np.min([np.max([0,.5*TempTarg]),3])])
        
        indexspeed = int(np.round(fanspeed))
        HPState[pump]['fanLevel'] = FanSpeeds[indexspeed]
    
    for pump in Pumps:
        print(pump+' last switch at : '+str(OnOffSwitch[pump]) )
 
    ###### HERE COMMANDS ARE SENT TO THE PUMPS, COMMENT OUT TO MAKE THE CODE "PASSIVE" ######
    print('Update pumps states')
    UpdatePumpStates(SensiboDevices,HPState)
    ########################################################################################
    
    # Plot data
    FN = 5
    # Temperatures
    Leg = ['Measured','Target','MHE']
    for pump in Pumps:
            AxList[FN][0].clear()
            Ax2List[FN][0].clear()
            #try:
            AxList[FN][0].step(DataHP[pump]['flattime'],DataHP[pump]['measurements']['temperature'], where='post',color=[.8,.8,1],linestyle = '-')
            AxList[FN][0].step(DataHP[pump]['flattime'],DataHP[pump]['states']['targetTemperature'], where='post',color='g',linestyle = '-')
            AxList[FN][0].step(MHELog['Times'],MHELog[pump]['State']['Temp'], where='post', linewidth=2,color='r',linestyle='-')
            Ax2List[FN][0].step(DataHP[pump]['flattime'],DataHP[pump]['states']['on'], where='post', linewidth=2,color='k',linestyle='-')
                    
            if MHEEstimationAvailable:
                AxList[FN][0].step(MHE_time_grid,wMHE_opt['State',:,pump,'Temp'],color='b',linewidth=2,where='post')
                Leg.append('MHE Sol.')
            if MPCControlAvailable:
                AxList[FN][0].step(MPC_time_grid,wMPC_opt['State',:,pump,'Temp'],color='b',linewidth=2,linestyle = '--',where='post')
                AxList[FN][0].step(MPC_time_grid,wMPC_opt['State',:,pump,'TargetTemp'],color='g',linewidth=2,linestyle = '--',where='post')
                Ax2List[FN][0].step(MPC_time_grid,wMPC_opt['State',:,pump,'Discomfort'],color='c',linewidth=2,linestyle = '--',where='post')
                #Ax2List[FN][0].step(MPC_time_grid[:-1],wMPC_opt['Input',:,pump,'Slack'],color='y',linewidth=2,linestyle = '--',where='post')

                AxList[FN][0].step(MPC_time_grid,DataMPCNum['DesiredTemperature',:,pump],color='m',linewidth=2,linestyle = '-',where='post')
                AxList[FN][0].step(MPC_time_grid,DataMPCNum['MinTemperature',:,pump],color='r',linewidth=2,linestyle = '-',where='post')
                Leg.append('MPC room temp')
                Leg.append('MPC temp setting')
                Leg.append('Desired temp.')
                Leg.append('Min temp.')
            AxList[FN][0].legend(Leg)
            Ax2List[FN][0].legend(['on/off','Discomfort'])
            #Ax2List[FN][0].legend('HP on')
            AxList[FN][0].grid()
            AxList[FN][0].set(title='Temp. in '+HPNames[pump])
            
            AxList[FN][1].clear()
            Ax2List[FN][1].clear()
            AxList[FN][1].set(title=HPNames[pump]+' power')
            if MPCControlAvailable:
                AxList[FN][1].step(MPC_time_grid,wMPC_opt['State',:,pump,'Power'], where='post', linewidth=2,color='b')
                AxList[FN][1].step(MPC_time_grid[:-1],wMPC_opt['Input',:,pump,'On'], where='post', linewidth=2,color='k')
                Ax2List[FN][1].step(MPC_time_grid,np.array(DataMPCNum['SpotPrices',:])+GridCost, where='post', linewidth=2,color='r')
                Ax2List[FN][1].set_ylabel('Prices [Øre]',fontsize=20,color='r')
                AxList[FN][1].set_ylabel('Power [kwh]',fontsize=20,color='b')
            AxList[FN][1].grid()
            FigList[FN].canvas.draw()
            FN += 1
    
    # Spot prices (Figure 10)
    AxList[FN][0].clear()
    Ax2List[FN][0].clear()
    try:
        #Ax2[0].step( PowerLog['Times'], PowerLog['Power'] ,color=[.8,.8,1])
        Ax2List[FN][0].step( time_grid, PowerAverage,color=[.8,.8,1])
        AxList[FN][0].step(Spot['Time'],np.array(Spot['Prices'])+GridCost, where='post', linewidth=2,color='r')
        AxList[FN][0].plot([TimeSchedule]*len(Spot['Prices']),np.array(Spot['Prices'])+GridCost,linewidth=2,color='k')
        AxList[FN][0].plot([TimeSchedule+timedelta(hours=MPC_Horizon)]*len(Spot['Prices']),np.array(Spot['Prices'])+GridCost,linewidth=1,color='k',linestyle='--')
        AxList[FN][0].text(TimeSchedule,max(Spot['Prices'])+GridCost,'Now',horizontalalignment='center',verticalalignment='bottom',fontsize=15)
        AxList[FN][0].plot([TimeSchedule, TimeSchedule+timedelta(hours=MPC_Horizon)],[DataMPCNum['BasePrice']]*2,linewidth=1,color='c',linestyle='-')
        Ax2List[FN][0].set_ylim([0,4.5])
        AxList[FN][0].set_ylim([np.min(np.array(DataMPCNum['SpotPrices',:])+GridCost),np.max(np.array(DataMPCNum['SpotPrices',:])+GridCost)])
    except:
        print('Spot prices plotting failed')
  
    AxList[FN][0].set(ylabel='Price (Øre/kWh)')
    AxList[FN][0].set(xlabel='Local time')
    AxList[FN][0].autoscale(enable=True, axis='x', tight=True)
    AxList[FN][0].autoscale(enable=True, axis='y', tight=True)
    AxList[FN][0].grid()
    FigList[FN].canvas.draw()

    AxList[FN][1].clear()
    Ax2List[FN][1].clear()
    try:
        Leg = ['Av.','MHE log']
        AxList[FN][1].step( time_grid, PowerAverage ,color=[.8,.8,1])
        AxList[FN][1].step(MHELog['Times'],MHELog['Power'], where='post', linewidth=2,color='k',linestyle='-')
        if MHEEstimationAvailable:
            AxList[FN][1].step(MHE_time_grid,MHEElecPower,color='b',linewidth=2,where='post')
            Leg.append('MHE Sol.')
        if MPCControlAvailable:
            AxList[FN][1].step(MPC_time_grid[:-1], MPCElecPower,color='b',linewidth=2,linestyle = '--',where='post')
            Leg.append('MPC Sol.')
        AxList[FN][1].set_ylabel('Power [kW]',fontsize=20,color='b')
        Ax2List[FN][1].step(MPC_time_grid,np.array(DataMPCNum['SpotPrices',:])+GridCost, where='post', linewidth=2,color='r')
        Ax2List[FN][1].set_ylabel('Total Price [Øre]',fontsize=20,color='r')
        Ax2List[FN][1].plot([TimeSchedule, TimeSchedule+timedelta(hours=MPC_Horizon)],[DataMPCNum['BasePrice']]*2,linewidth=1,color='c',linestyle='-')
        AxList[FN][1].legend(Leg)
        AxList[FN][1].set_xlim([MHE_time_grid[0],MPC_time_grid[-1]])
        AxList[FN][1].set_ylim([0,4.5])
        Ax2List[FN][1].set_ylim([np.min(np.array(DataMPCNum['SpotPrices',:])+GridCost),np.max(np.array(DataMPCNum['SpotPrices',:])+GridCost)])
        AxList[FN][1].grid()
    except:
        print('Real-time energy plot failed')
        sys.exit()

        
    FigList[FN].canvas.draw()
    FN += 1
    
    # Weather
    try:
        AxList[FN][0].clear()
        AxList[FN][0].step(Weather['Time'],Weather['Temperature'],color='b')
        AxList[FN][0].grid()
        AxList[FN][1].clear()
        Ax2List[FN][1].clear()
        AxList[FN][1].step(Forecast['Time'],Forecast['Temperature'],color='c')
        Ax2List[FN][1].step(RadiationData['Time'],RadiationData['DNI'],color='c')
        AxList[FN][1].grid()
    except:
        print('Weather plot failed')
    
    FigList[FN].canvas.draw()
    
    FN += 1
    
    try:
        AxList[FN][0].clear()
        AxList[FN][0].step(Weather['Time'],Weather['Temperature'],color='b')
        AxList[FN][0].grid()
        AxList[FN][1].clear()
        AxList[FN][1].step(RadiationData['Time'],RadiationData['DNI'],color='c')
        AxList[FN][1].plot([TimeSchedule]*len(Radiation['DNI']),Radiation['DNI'],linewidth=2,color='k')
        AxList[FN][1].grid()
        FigList[FN].canvas.draw()
    except:
        print('Radiation plot failed')
    
    FigList[FN].canvas.draw()
        
    #### PLOTS TO ASSESS MPC / MHE PREDICTION / FITTING ####
    FN = 12

    NMPCtraj = int(2*60/float(SamplingTime['Measurement']))
    NDisp    = np.min([len(Logger_MHE_MPC['MPC']['RawSolution']),NMPCtraj])
    NMPCdisphorizon = int(12*60/float(SamplingTime['Measurement']))
    LenLogger = len(Logger_MHE_MPC['MPC']['RawSolution'])
    for fig, pump in enumerate(Pumps):
        AxList[FN][0].clear()
        for index in range(0,NDisp): #, traj in enumerate(Logger_MHE_MPC['MPC']['RawSolution'][-NMPCtraj:]):
            traj = Logger_MHE_MPC['MPC']['RawSolution'][index + LenLogger - NDisp]
            Col  = [(NDisp-index)/float(NDisp),(NDisp-index)/float(NDisp),(NDisp-index)/float(NDisp)]
            Traj = wMPC(traj)
            AxList[FN][0].step(Logger_MHE_MPC['MPC']['Time_Grid'][index-NDisp][:NMPCdisphorizon],Traj['State',:NMPCdisphorizon,pump,'Temp'], where='post', linewidth=1,color=Col,linestyle='-')
            AxList[FN][0].step(Logger_MHE_MPC['MPC']['Time_Grid'][index-NDisp][0],Traj['State',0,pump,'Temp'], where='post', marker='.',color='r')
        AxList[FN][0].step(Logger_MHE_MPC['MPC']['Time_Grid'][-1][:NMPCdisphorizon],Traj['State',:NMPCdisphorizon,pump,'Temp'], where='post', linewidth=1,color=[.5,0,0],linestyle='-')
        AxList[FN][0].step(MHELog['Times'],MHELog[pump]['State']['Temp'], where='post', linewidth=2,color=[.35,0,0],linestyle='-')
        AxList[FN][0].step(DataHP[pump]['flattime'],DataHP[pump]['measurements']['temperature'], where='post', linewidth=2,color='b',linestyle = '-')
        
        AxList[FN][1].clear()
        for index in range(0,NDisp): #traj in enumerate(Logger_MHE_MPC['MPC']['RawSolution'][-NMPCtraj:]):
            traj = Logger_MHE_MPC['MPC']['RawSolution'][index + LenLogger - NDisp]
            Traj = wMPC(traj)
            Col  = [(NDisp-index)/float(NDisp),(NDisp-index)/float(NDisp),(NDisp-index)/float(NDisp)]
            AxList[FN][1].step(Logger_MHE_MPC['MPC']['Time_Grid'][index-NDisp][:NMPCdisphorizon],Traj['State',:NMPCdisphorizon,pump,'TargetTemp'], where='post', linewidth=1,color=Col,linestyle='-')
        AxList[FN][1].step(Logger_MHE_MPC['MPC']['Time_Grid'][-1][:NMPCdisphorizon],Traj['State',:NMPCdisphorizon,pump,'TargetTemp'], where='post', linewidth=1,color=[.5,0,0],linestyle='-')
        AxList[FN][1].step(DataHP[pump]['flattime'],DataHP[pump]['states']['targetTemperature'], where='post', linewidth=2,color='g',linestyle = '-')

        AxList[FN][0].set(title=pump+' temp.')
        AxList[FN][1].set(title=pump+' target temp.')
        AxList[FN][0].grid()
        AxList[FN][1].grid()
        FigList[FN].canvas.draw()
        FN += 1

    FN = 16
    AxList[FN][0].clear()
    Ax2List[FN][0].clear()
    AxList[FN][0].step(TibberData['flattime'],np.array(TibberData['Power'])*1e-3, where='post', linewidth=1,color=[1,.9,.9],linestyle='-')
    AxList[FN][0].step(time_grid,PowerAverage, where='post', linewidth=1,color='r',linestyle='-')
    for index in range(0,NDisp): #for index, traj in enumerate(Logger_MHE_MPC['MPC']['RawSolution'][-NMPCtraj:]):
        traj = Logger_MHE_MPC['MPC']['RawSolution'][index + LenLogger - NDisp]
        Col  = [(NDisp-index)/float(NDisp),(NDisp-index)/float(NDisp),(NDisp-index)/float(NDisp)]
        PredPower = 0
        for pump in Pumps:
            Traj = wMPC(traj)
            PredPower += np.array(Traj['Input',:NMPCdisphorizon,pump,'On'])*np.array(Traj['State',:NMPCdisphorizon,pump,'Power'])
            
        AxList[FN][0].step(Logger_MHE_MPC['MPC']['Time_Grid'][index-NDisp][:NMPCdisphorizon],PredPower, where='post', linewidth=1,color=Col,linestyle='-')
    AxList[FN][0].step(MHE_time_grid,MHEElecPower,color='c',linewidth=2,where='post')

    AxList[FN][0].step(MHELog['Times'],MHELog['Power'], where='post', linewidth=2,color='b',linestyle='-')
    Ax2List[FN][0].step(SpotMarket['flattime'],np.array(SpotMarket['Prices'])+GridCost, where='post', linewidth=2,color=[.5,0,.5])
    Ax2List[FN][0].plot(MPC_time_grid,np.array(DataMPCNum['SpotPrices',:])+GridCost, marker='o', linewidth=1,color='m')
    Ax2List[FN][0].set_ylabel('Prices [Øre]',fontsize=20,color='m')
    AxList[FN][0].set(title='Total Power')
    AxList[FN][0].set_ylim([0,4.5])
    Ax2List[FN][0].set_ylim([np.min(DataMPCNum['SpotPrices',:])+GridCost,np.max(DataMPCNum['SpotPrices',:])+GridCost])
    
    Ax2List[FN][0].set_ylim([np.min(SpotMarket['Prices'])+GridCost,np.max(np.array(SpotMarket['Prices']))+GridCost])
    
    AxList[FN][1].clear()
    Ax2List[FN][1].clear()
    for index in range(0,NDisp): #for index, traj in enumerate(Logger_MHE_MPC['MPC']['RawSolution'][-NMPCtraj:]):
        traj = Logger_MHE_MPC['MPC']['RawSolution'][index + LenLogger - NDisp]
        Col  = [(NDisp-index)/float(NDisp),(NDisp-index)/float(NDisp),(NDisp-index)/float(NDisp)]
        PredPower = 0
        for pump in Pumps:
            Traj = wMPC(traj)
            PredPower += np.array(Traj['Input',:NMPCdisphorizon,pump,'On'])*np.array(Traj['State',:NMPCdisphorizon,pump,'Power'])
            
        AxList[FN][1].step(Logger_MHE_MPC['MPC']['Time_Grid'][index-NDisp][:NMPCdisphorizon],PredPower, where='post', linewidth=1,color=Col,linestyle='-')
    AxList[FN][1].step(time_grid,PowerAverage, where='post', linewidth=1,color='r',linestyle='-')
    AxList[FN][1].step(MHE_time_grid,MHEElecPower,color='c',linewidth=2,where='post')

    AxList[FN][1].step(MHELog['Times'],MHELog['Power'], where='post', linewidth=2,color='b',linestyle='-')
    Ax2List[FN][1].step(MPC_time_grid,np.array(DataMPCNum['SpotPrices',:])+GridCost, where='post', linewidth=2,color='m')
    Ax2List[FN][1].set_ylabel('Prices [Øre]',fontsize=20,color='m')
    AxList[FN][1].set(title='Total Power')
    AxList[FN][1].set_ylim([0,4.5])
    Ax2List[FN][1].set_ylim([np.min(DataMPCNum['SpotPrices',:])+GridCost,np.max(DataMPCNum['SpotPrices',:])+GridCost])

    FigList[FN].canvas.draw()
    
    FN = 17
    Col = ['r','g','b','m']
    for sub, key in enumerate(SYSID[Pumps[0]].keys()):
        AxList[FN][sub].clear()
        Leg = []
        for col, pump in enumerate(Pumps):
            AxList[FN][sub].step(SYSIDLog['Time'],SYSIDLog[pump][key],color=Col[col],where='post')
            #AxList[FN][sub].step(SYSIDLog['Time'],[SYSID[pump][key]]*len(SYSIDLog['Time']),color=Col[col],linestyle=':',where='post')
            Leg.append(pump)
        AxList[FN][sub].set(title=key)
        AxList[FN][0].legend(Leg)
    FigList[FN].canvas.draw()
    
    # Plot all
    plt.pause(0.1) #hack to ensure that the plot is drawn on the fly
    plt.show(block=False)
    
    ## Prepare next measurement time
    time          = datetime.now(tz=local_timezone)
    TimeSchedule += timedelta(minutes=SamplingTime['Measurement'])
    
    # If code stalled and next measurement is in the past, iterate until it's in the future.
    while TimeSchedule < time:
        print('Pushed time schedule forward')
        TimeSchedule += timedelta(minutes=SamplingTime['Measurement'])



    DataPickle = {'TimeGrid'          : time_grid,
                  'Times'             : Times,
                  'TimeInitial'       : TimeInitial,
                  'Sampling'          : SamplingTime,
                  'EnergyRT'          : PowerLog,
                  'Power'             : PowerAverage,
                  'HP'                : DataHP,
                  'Weather'           : Weather,
                  'Radiation'         : RadiationData,
                  'RadiationForecast' : Radiation,
                  'LoggerMHEMPC'      : Logger_MHE_MPC,
                  'WeightsLog'        : WeightsLog,
                  'SYSIDLog'          : SYSIDLog
            }

    f = open(FileName+'.pkl',"wb")
    pickle.dump(DataPickle,f, protocol=2)
    f.close()



